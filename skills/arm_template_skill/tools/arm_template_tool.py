"""ARM Template Analyser — standalone CLI tool for Azure resource inventory extraction."""

import json
from pathlib import Path
from typing import Dict, List, Optional, Callable, Tuple, Set
from datetime import datetime
from collections import defaultdict
import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class ArmTemplateAnalyser:
    """Analyzes Azure ARM template JSON and extracts resource inventory."""

    SEVERITY_COLORS = {
        "CRITICAL": "FF0000",  # red
        "HIGH": "FF6600",  # orange
        "MEDIUM": "FFFF00",  # yellow
        "LOW": "CCCCCC",  # grey
    }

    def __init__(
        self,
        input_path: str,
        output_path: str,
        progress_callback: Optional[Callable] = None,
        verbose: bool = False,
    ):
        """Initialize analyser."""
        self.input_path = Path(input_path)
        self.output_path = Path(output_path)
        self.callback = progress_callback or (print if verbose else lambda m: None)

        self.template: Dict = {}
        self.resources: List[Dict] = []
        self.parameters: Dict = {}
        self.variables: Dict = {}
        self.outputs: Dict = {}
        self.dependency_map: List[Dict] = []
        self.risk_flags: List[Dict] = []

    def run(self) -> Dict:
        """Execute analysis pipeline and return results."""
        try:
            self._load_template()
            self._extract_resources()
            self._extract_metadata()
            self._build_dependency_map()
            self._detect_risks()

            # Write outputs
            self.output_path.mkdir(parents=True, exist_ok=True)
            self._write_excel()
            self._write_json_summary()
            summary_md = self._write_markdown_report()

            return {
                "summary": summary_md,
                "output_files": [
                    str(self.output_path / "inventory.xlsx"),
                    str(self.output_path / "inventory_summary.json"),
                    str(self.output_path / "report.md"),
                ],
                "data": {
                    "resource_count": len(self.resources),
                    "resource_types": self._get_resource_type_counts(),
                    "risk_flags": self.risk_flags,
                    "resources": self.resources,
                    "dependencies": self.dependency_map,
                    "parameters": list(self.parameters.keys()),
                    "outputs": list(self.outputs.keys()),
                },
            }
        except Exception as e:
            return {
                "summary": f"## Error\n\n{str(e)}",
                "output_files": [],
                "data": {"error": str(e)},
            }

    def _load_template(self):
        """Load and validate ARM template JSON."""
        self.callback("Reading ARM template...")
        if not self.input_path.exists():
            raise FileNotFoundError(f"Template file not found: {self.input_path}")

        try:
            self.template = json.loads(self.input_path.read_text())
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in template: {e}")

        # Validate ARM template
        if "$schema" not in self.template:
            raise ValueError("Template missing $schema — not a valid ARM template")

        schema = self.template.get("$schema", "")
        if "deploymentTemplate" not in schema:
            raise ValueError(f"Unknown schema: {schema}")

    def _extract_resources(self):
        """Extract resources from template with detailed properties."""
        self.callback("Extracting resources...")
        resources_arr = self.template.get("resources", [])

        for res in resources_arr:
            self._process_resource(res, parent_resource=None)

    def _process_resource(self, res: Dict, parent_resource: Optional[str] = None):
        """Process a single resource, handling child resources."""
        name = res.get("name", "")
        res_type = res.get("type", "")
        api_version = res.get("apiVersion", "")
        location = res.get("location", "")
        kind = res.get("kind", "")
        tags = res.get("tags", {})
        depends_on = res.get("dependsOn", [])
        properties = res.get("properties", {})
        sku = res.get("sku", {})

        # Parse provider and resource kind
        parts = res_type.split("/")
        provider = parts[0] if parts else ""
        resource_kind = parts[-1] if parts else ""

        # Extract key properties for common resource types
        extracted_props = self._extract_key_properties(res_type, properties)

        resource_entry = {
            "name": name,
            "type": res_type,
            "provider": provider,
            "resource_kind": resource_kind,
            "api_version": api_version,
            "location": location,
            "kind": kind,
            "sku_name": sku.get("name") if sku else None,
            "sku_tier": sku.get("tier") if sku else None,
            "sku_capacity": sku.get("capacity") if sku else None,
            "tag_count": len(tags) if tags else 0,
            "tags": json.dumps(tags) if tags else "{}",
            "depends_on": json.dumps(depends_on) if depends_on else "[]",
            "depends_on_count": len(depends_on) if depends_on else 0,
            "property_keys": ",".join(properties.keys()) if properties else "",
            "child_resources": len(res.get("resources", [])),
            "parent_resource": parent_resource,
            **extracted_props,
        }

        self.resources.append(resource_entry)

        # Process child resources
        for child in res.get("resources", []):
            self._process_resource(child, parent_resource=name)

    def _extract_key_properties(self, res_type: str, properties: Dict) -> Dict:
        """Extract specific properties based on resource type."""
        extracted = {}

        if "storageAccounts" in res_type:
            extracted["accessTier"] = properties.get("accessTier")
            extracted["supportsHttpsTrafficOnly"] = properties.get(
                "supportsHttpsTrafficOnly"
            )
            extracted["minimumTlsVersion"] = properties.get("minimumTlsVersion")
            extracted["allowBlobPublicAccess"] = properties.get("allowBlobPublicAccess")

        elif "virtualMachines" in res_type:
            hw_profile = properties.get("hardwareProfile", {})
            os_profile = properties.get("osProfile", {})
            storage_profile = properties.get("storageProfile", {})
            extracted["vmSize"] = hw_profile.get("vmSize")
            extracted["computerName"] = os_profile.get("computerName")
            extracted["adminUsername"] = os_profile.get("adminUsername")
            os_disk = storage_profile.get("osDisk", {})
            managed_disk = os_disk.get("managedDisk", {})
            extracted["osDiskStorageType"] = managed_disk.get("storageAccountType")

        elif "virtualNetworks" in res_type:
            addr_space = properties.get("addressSpace", {})
            extracted["addressPrefixes"] = ",".join(
                addr_space.get("addressPrefixes", [])
            )
            subnets = properties.get("subnets", [])
            extracted["subnetCount"] = len(subnets)
            dns_servers = properties.get("dhcpOptions", {}).get("dnsServers", [])
            extracted["dnsServers"] = ",".join(dns_servers)

        elif "networkSecurityGroups" in res_type:
            rules = properties.get("securityRules", [])
            extracted["ruleCount"] = len(rules)
            inbound = sum(1 for r in rules if r.get("direction") == "Inbound")
            outbound = sum(1 for r in rules if r.get("direction") == "Outbound")
            extracted["inboundRules"] = inbound
            extracted["outboundRules"] = outbound

        elif "vaults" in res_type:
            extracted["enableSoftDelete"] = properties.get("enableSoftDelete")
            extracted["enablePurgeProtection"] = properties.get(
                "enablePurgeProtection"
            )
            access_policies = properties.get("accessPolicies", [])
            extracted["accessPolicieCount"] = len(access_policies)

        return extracted

    def _extract_metadata(self):
        """Extract parameters, variables, and outputs."""
        self.callback("Extracting parameters, variables and outputs...")
        self.parameters = self.template.get("parameters", {})
        self.variables = self.template.get("variables", {})
        self.outputs = self.template.get("outputs", {})

    def _build_dependency_map(self):
        """Build resource dependency map and detect cycles."""
        self.callback("Building dependency map...")
        self.dependency_map = []

        for res in self.resources:
            depends_on_raw = json.loads(res["depends_on"])
            for dep in depends_on_raw:
                # Extract resource name from resourceId() or reference expressions
                dep_name = self._extract_resource_name(dep)
                self.dependency_map.append(
                    {
                        "resource": res["name"],
                        "depends_on": dep,
                        "resolved_name": dep_name,
                    }
                )

        # Detect circular dependencies
        self._detect_circular_dependencies()

    def _extract_resource_name(self, dep_str: str) -> str:
        """Extract resource name from dependency string."""
        # Handle resourceId(type, name) format
        if "resourceId" in dep_str:
            match = re.search(r"'([^']+)'", dep_str)
            if match:
                return match.group(1)
        return dep_str.strip()

    def _detect_circular_dependencies(self):
        """Detect and flag circular dependencies."""
        # Build adjacency map
        graph = defaultdict(list)
        for dep in self.dependency_map:
            graph[dep["resource"]].append(dep["resolved_name"])

        # DFS to find cycles
        visited = set()
        rec_stack = set()

        def dfs(node: str, path: List[str]) -> bool:
            visited.add(node)
            rec_stack.add(node)
            path.append(node)

            for neighbor in graph.get(node, []):
                if neighbor not in visited:
                    if dfs(neighbor, path):
                        return True
                elif neighbor in rec_stack:
                    # Found cycle
                    cycle_start = path.index(neighbor) if neighbor in path else 0
                    cycle = path[cycle_start:] + [neighbor]
                    self.risk_flags.append(
                        {
                            "resource": node,
                            "type": "dependency",
                            "flag": "Circular dependency",
                            "severity": "CRITICAL",
                            "detail": " → ".join(cycle),
                        }
                    )
                    return True

            rec_stack.discard(node)
            path.pop()
            return False

        for node in graph:
            if node not in visited:
                dfs(node, [])

    def _detect_risks(self):
        """Detect risk flags across resources."""
        self.callback("Detecting risk flags...")

        for res in self.resources:
            # CRITICAL: circular dependencies (already detected)

            # HIGH: missing apiVersion
            if not res["api_version"]:
                self.risk_flags.append(
                    {
                        "resource": res["name"],
                        "type": res["type"],
                        "flag": "Missing apiVersion",
                        "severity": "HIGH",
                        "detail": "Resource missing explicit API version",
                    }
                )

            # HIGH: storage with public blob access enabled or not set
            if "storageAccounts" in res["type"]:
                if res.get("allowBlobPublicAccess") in (True, "true", "True", None):
                    self.risk_flags.append(
                        {
                            "resource": res["name"],
                            "type": res["type"],
                            "flag": "Public blob access not restricted",
                            "severity": "HIGH",
                            "detail": "allowBlobPublicAccess is enabled or not set",
                        }
                    )

            # HIGH: VM without managed disk
            if "virtualMachines" in res["type"]:
                if not res.get("osDiskStorageType"):
                    self.risk_flags.append(
                        {
                            "resource": res["name"],
                            "type": res["type"],
                            "flag": "No managed disk defined",
                            "severity": "HIGH",
                            "detail": "VM lacks managed disk for OS",
                        }
                    )

            # MEDIUM: missing location
            if not res["location"]:
                self.risk_flags.append(
                    {
                        "resource": res["name"],
                        "type": res["type"],
                        "flag": "Missing location",
                        "severity": "MEDIUM",
                        "detail": "Resource has no explicit location",
                    }
                )

            # MEDIUM: apiVersion older than 2020-01-01
            if res["api_version"] and res["api_version"] < "2020-01-01":
                self.risk_flags.append(
                    {
                        "resource": res["name"],
                        "type": res["type"],
                        "flag": "Outdated API version",
                        "severity": "MEDIUM",
                        "detail": f"Using {res['api_version']} (older than 2020-01-01)",
                    }
                )

            # MEDIUM: KeyVault without soft delete
            if "vaults" in res["type"]:
                if res.get("enableSoftDelete") not in (True, "true", "True"):
                    self.risk_flags.append(
                        {
                            "resource": res["name"],
                            "type": res["type"],
                            "flag": "Soft delete not enabled",
                            "severity": "MEDIUM",
                            "detail": "KeyVault should enable soft delete protection",
                        }
                    )

            # LOW: resource with no tags
            if res["tag_count"] == 0:
                self.risk_flags.append(
                    {
                        "resource": res["name"],
                        "type": res["type"],
                        "flag": "No tags defined",
                        "severity": "LOW",
                        "detail": "Resource has no tags for organization/billing",
                    }
                )

            # LOW: no SKU defined (for resources that typically need one)
            if not res["sku_name"] and any(
                t in res["type"]
                for t in [
                    "appServicePlans",
                    "databases",
                    "elasticPools",
                    "registries",
                ]
            ):
                self.risk_flags.append(
                    {
                        "resource": res["name"],
                        "type": res["type"],
                        "flag": "SKU not defined",
                        "severity": "LOW",
                        "detail": "Resource typically requires explicit SKU",
                    }
                )

    def _get_resource_type_counts(self) -> Dict[str, int]:
        """Count resources by type."""
        counts = defaultdict(int)
        for res in self.resources:
            counts[res["type"]] += 1
        return dict(counts)

    def _write_excel(self):
        """Write professional-style Excel inventory."""
        self.callback("Writing inventory.xlsx...")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define styles
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Sheet 1: All Resources
        ws = wb.create_sheet("All Resources")
        headers = [
            "Name",
            "Type",
            "Provider",
            "API Version",
            "Location",
            "Kind",
            "SKU Name",
            "Tags",
            "Depends On Count",
            "Child Resources",
        ]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_num, res in enumerate(self.resources, 2):
            values = [
                res["name"],
                res["type"],
                res["provider"],
                res["api_version"],
                res["location"],
                res["kind"],
                res["sku_name"],
                res["tags"],
                res["depends_on_count"],
                res["child_resources"],
            ]
            ws.append(values)

            fill = alt_fill if row_num % 2 == 0 else PatternFill()
            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Auto-filter and freeze pane
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

        # Auto-width columns
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = min(45, max(12, len(header) + 2))

        # Sheet 2+: Resources by type
        for res_type in sorted(self._get_resource_type_counts().keys()):
            # Clean name for sheet (max 31 chars)
            sheet_name = res_type.replace("Microsoft.", "").replace("/", " ")[:31]
            ws = wb.create_sheet(sheet_name)

            type_resources = [r for r in self.resources if r["type"] == res_type]
            if type_resources:
                headers = [
                    "Name",
                    "API Version",
                    "Location",
                    "Kind",
                    "Tags",
                    "Depends On Count",
                ]
                ws.append(headers)

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border

                for row_num, res in enumerate(type_resources, 2):
                    values = [
                        res["name"],
                        res["api_version"],
                        res["location"],
                        res["kind"],
                        res["tags"],
                        res["depends_on_count"],
                    ]
                    ws.append(values)

                    for col_num, val in enumerate(values, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)

                ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # Sheet: Parameters
        ws = wb.create_sheet("Parameters")
        headers = ["Name", "Type", "Default Value", "Description"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_num, (name, param) in enumerate(self.parameters.items(), 2):
            values = [
                name,
                param.get("type", ""),
                json.dumps(param.get("defaultValue", "")),
                param.get("metadata", {}).get("description", ""),
            ]
            ws.append(values)

            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border

        # Sheet: Risk Flags
        ws = wb.create_sheet("Risk Flags")
        headers = ["Severity", "Resource", "Type", "Flag", "Detail"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_num, flag in enumerate(self.risk_flags, 2):
            severity = flag["severity"]
            color = self.SEVERITY_COLORS.get(severity, "CCCCCC")
            sev_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            values = [
                flag["severity"],
                flag["resource"],
                flag["type"],
                flag["flag"],
                flag["detail"],
            ]
            ws.append(values)

            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if col_num == 1:
                    cell.fill = sev_fill
                cell.alignment = Alignment(wrap_text=True)

        wb.save(self.output_path / "inventory.xlsx")

    def _write_json_summary(self):
        """Write structured JSON summary."""
        self.callback("Writing inventory_summary.json...")

        summary = {
            "template_file": self.input_path.name,
            "analysed_at": datetime.utcnow().isoformat(),
            "schema_version": self.template.get("$schema", "unknown"),
            "content_version": self.template.get("contentVersion", "unknown"),
            "resource_count": len(self.resources),
            "resource_types": self._get_resource_type_counts(),
            "parameters": list(self.parameters.keys()),
            "variables": list(self.variables.keys()),
            "outputs": list(self.outputs.keys()),
            "risk_summary": self._get_risk_summary(),
            "risk_flags": self.risk_flags,
            "resources": self.resources,
            "dependencies": self.dependency_map,
        }

        (self.output_path / "inventory_summary.json").write_text(
            json.dumps(summary, indent=2)
        )

    def _get_risk_summary(self) -> Dict[str, int]:
        """Summarize risks by severity."""
        summary = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
        for flag in self.risk_flags:
            summary[flag["severity"]] += 1
        return summary

    def _write_markdown_report(self) -> str:
        """Write markdown report and return its content."""
        self.callback("Generating markdown report...")

        lines = [
            "# ARM Template Analysis Report",
            f"\n**File:** `{self.input_path.name}`",
            f"**Analysed:** {datetime.utcnow().isoformat()}",
            f"**Schema:** {self.template.get('$schema', 'unknown')}",
            "\n## Summary",
            "\n| Metric | Value |",
            "|---|---|",
        ]

        risk_summary = self._get_risk_summary()
        lines.extend(
            [
                f"| Total resources | {len(self.resources)} |",
                f"| Resource types | {len(self._get_resource_type_counts())} |",
                f"| Parameters | {len(self.parameters)} |",
                f"| Outputs | {len(self.outputs)} |",
                f"| Risk flags | {len(self.risk_flags)} |",
                f"| CRITICAL risks | {risk_summary['CRITICAL']} |",
                f"| HIGH risks | {risk_summary['HIGH']} |",
            ]
        )

        lines.extend(
            [
                "\n## Resources by type",
                "\n| Type | Count |",
                "|---|---|",
            ]
        )
        for res_type, count in sorted(self._get_resource_type_counts().items()):
            lines.append(f"| {res_type} | {count} |")

        # Risk flags
        lines.append("\n## Risk flags")

        if self.risk_flags:
            for severity in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
                severity_flags = [f for f in self.risk_flags if f["severity"] == severity]
                if severity_flags:
                    lines.append(f"\n### {severity}")
                    for flag in severity_flags:
                        lines.append(
                            f"\n- **{flag['resource']}** (`{flag['type']}`): {flag['flag']}"
                        )
                        lines.append(f"  - {flag['detail']}")
        else:
            lines.append("\nNo risk flags detected.")

        # Dependencies
        if self.dependency_map:
            lines.extend(
                [
                    "\n## Dependency map",
                    "\n| Resource | Depends On |",
                    "|---|---|",
                ]
            )
            for dep in self.dependency_map:
                lines.append(f"| {dep['resource']} | {dep['depends_on']} |")

        # Parameters
        if self.parameters:
            lines.extend(
                [
                    "\n## Parameters",
                    "\n| Name | Type | Default |",
                    "|---|---|---|",
                ]
            )
            for name, param in self.parameters.items():
                param_type = param.get("type", "")
                default = param.get("defaultValue", "—")
                lines.append(f"| {name} | {param_type} | {default} |")

        # Output files
        lines.extend(
            [
                "\n## Output files",
                "\n- `inventory.xlsx` — full resource table with risk sheet",
                "\n- `inventory_summary.json` — structured JSON data",
                "\n- `report.md` — this report",
            ]
        )

        report_md = "\n".join(lines)
        (self.output_path / "report.md").write_text(report_md)

        return report_md


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="ARM Template Analyser")
    parser.add_argument("--input", required=True, help="Path to ARM template JSON")
    parser.add_argument("--output", required=True, help="Output folder")
    parser.add_argument("--verbose", action="store_true", help="Verbose output")
    args = parser.parse_args()

    analyser = ArmTemplateAnalyser(args.input, args.output, verbose=args.verbose)
    result = analyser.run()
    print(result["summary"])
